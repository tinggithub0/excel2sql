import openpyxl
from top_user_id import setting
from check_fun import check_gamelist, check_dictionary, check_gamemenu
from convert_fun import convert_gamelist, convert_dictionary, convert_gamemenu


wb = openpyxl.load_workbook('123.xlsx')
top_user_id = setting.top_user_id
list_content = {}
dict_content = {}
menu_content = {}
gamelist_type = len(wb.sheetnames)


def read_cell(dict):
    for row_index, row in enumerate(sheet.rows):
        if row[0].value is None: # 空行跳過
            pass
        else:
            cells = []
            for cell in row:
                cells.append(cell.value)
                dict[row_index] = cells
    del dict[0]

def get_content(sheet):
    mark = ''
    if gamelist_type == 3: # 3 頁型
        if (sheet.cell(1,1).value == 'GameID') and (list_content == {}):
            read_cell(list_content)
            if check_gamelist(list_content) is False:
                print(f'錯誤 - 3-1, mark= "{mark}"')
            mark += '1'
        elif (sheet.cell(1,1).value == 'DictionaryType') and (dict_content == {}):
            read_cell(dict_content)
            if check_dictionary(dict_content) is False:
                print(f'錯誤 - 3-2, mark= "{mark}"')
            mark += '2'
        elif (sheet.cell(1,1).value == 'GameCode') and (menu_content == {}):
            read_cell(menu_content)
            if check_gamemenu(menu_content) is False:
                print(f'錯誤 - 3-3, mark= "{mark}"')
            mark += '3'
        else:
            print(f'錯誤 - 3, mark= "{mark}"')
            exit()
    elif gamelist_type == 1: # 1 頁型
        if (sheet.cell(1,1).value == 'DictionaryType') and (dict_content == {}):
            read_cell(dict_content)
            if check_dictionary(dict_content) is False:
                print(f'錯誤 - 1-1, mark= "{mark}"')
            mark += '4'
        else:
            print(f'錯誤 - 1, mark= "{mark}"')
            exit()


if __name__ == "__main__":
    for no in range(gamelist_type):
        sheet = wb.worksheets[no]
        get_content(sheet)
        with open('gamelist.txt', 'w', encoding='UTF-8') as f:
            if gamelist_type == 3:
                f.write("# GameList\n")
                convert_gamelist(f, list_content)
                f.write("\n\n\n# GameDictionary\n")
                convert_dictionary(f, dict_content)
                f.write("\n\n\n# GameMenu\n")
                convert_gamemenu(f, menu_content, top_user_id)
            elif gamelist_type == 1:
                f.write("# GameDictionary\n")
                convert_dictionary(f, dict_content)
